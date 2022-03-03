
import openpyxl

def read_data(excl_dir, sheet_name):
    """读取测试数据，通过Excel"""
    work_book = openpyxl.load_workbook(excl_dir, sheet_name)
    worksheet = work_book[sheet_name]
    list_new1 = []
    for i in range(2, int(worksheet.max_row)+1):
        list_new2 = []
        for j in range(1,worksheet.max_column+1):
            cell_data = worksheet.cell(i,j).value
            list_new2.append(cell_data)
        list_new1.append(list_new2)

    return list_new1







if __name__ == '__main__':
    excl_dir = r"../data/data.xlsx"
    sheet_name = "register"
    read_data(excl_dir, sheet_name)